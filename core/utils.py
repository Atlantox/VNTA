from openpyxl import load_workbook

from core.ActionChain import ActionChain
from core.Decision import Decision, Option

def GetFileNameAndFormat(path:str):
    ''' Get a file path and returns the file name without format, and the file format '''
    fileFullName = path.split('/')[-1]  # File name with format
    splits = fileFullName.split('.')
    fileFormat = '.' + splits[-1]

    fileName = ''
    for split in splits[:-1]:
        fileName += split
    
    return fileName, fileFormat


def ReadDecisions(filePath:str, fileFormat:str):
    ''' Return all decisions in the specified .csv or excel file and the novel points '''
    decisions = []
    if fileFormat == '.xlsx': # Excel file
        decisions, novel_points, endings = LoadDecisionsFromExcel(filePath)        
    elif fileFormat == '.csv':  # CSV file
        decisions, novel_points, endings = LoadDecisionsFromCSV(filePath)
        
    return decisions, novel_points, endings


def LoadDecisionsFromExcel(filePath:str):
    '''
    Return the Decisions list and novel points of target Excel file
    IMPORTANT: the decisions must stay in a sheet named 'decisions'
    '''
    decisions = []
    endings = {}
    excel = load_workbook(filename=filePath, read_only=True)
    sheet = excel['decisions']
    rows = sheet.iter_rows()
    all_rows = [r for r in rows if str(r[0].value) != 'None']

    i = 0
    while(i < len(all_rows)):
        row = all_rows[i]
        if i == 0:
            # Getting the novels points names
            novel_points = [str(c.value).strip() for c in row[5:] if str(c.value).strip() != '' and c.value is not None]    
        else:
            decision_name = str(row[2].value)
            related_rows = [row]
            relatedCounter = 1
            searching = True
            while(searching):
                try:
                    next_row = all_rows[i + relatedCounter]
                    # If the name is None and the id cell is not empty, then is a decision with more than one option

                    if str(next_row[2].value) == 'None':
                        related_rows.append(next_row)
                        relatedCounter += 1
                    else:
                        searching = False
                except:
                    searching = False
            
            i += relatedCounter - 1
            options = []
            for related in related_rows:
                #  Getting the first 5 fields (id, type, name, option, dependency)
                id, dtype, name, option, dependencies = [str(c.value) for c in related[:5]]
                if 'E-' == dtype[0:2]:
                    ending_name = dtype[2:]
                    if ending_name not in endings: endings[ending_name] = 0

                # The rest of fields are novel points
                points = [str(c.value) for c in related[5:5 + len(novel_points)]]
                points, dependencies = GetPointsAndDependencies(points, dependencies)

                if dtype == 'I' or 'E-' == dtype[0:2]:
                    if ',' in option: option = [n.strip() for n in option.split(',')]

                options.append(Option(
                    id=id, 
                    name=option,
                    dependencies=dependencies,
                    points=points
                ))

            # Once we have all related Options, create the Decision
            decisions.append(Decision(
                id=id,
                type=dtype,
                name=decision_name,
                options=options,
            ))

        i+=1

    return decisions, novel_points, endings


def GetPointsAndDependencies(points:list[str], dependencies:str):
    '''
    Obtain the points as list[str] and dependencies as str
    Return the points an dependencies ordered
    '''
    local_points = []
    local_dependencies = []
    for point in points:
        #  Converting the points in to useful values
        try:
            local_points.append(int(point))
        except:
            local_points.append(None)            

    if dependencies != 'None':
        # If the decision has dependency, convert the string into a list
        local_dependencies = [s.strip() for s in dependencies.split(',') if s != ''] 
    else: local_dependencies = None

    return local_points, local_dependencies


def LoadDecisionsFromCSV(filePath:str):
    ''' Return a list of Decisions and novel_points of target CSV file '''
    decisions = []
    endings = {}
    with open(filePath, 'r', encoding='utf-8') as f: #  Open the CSV file
        lines = f.readlines()
        f.close()
    
    # Getting the novel points names
    novel_points = [l.strip() for l in lines[0].split(';')[5:]]

    crt_id = 1
    rows = len(lines[1:])
    related_lines = []

    while crt_id < rows:
        # We get the first line
        current_line = lines[crt_id].split(';')
        related_lines = [current_line]
        
        count = 1
        while True:
            try:
                # If the index doesn't exists, then limit reached
                new_line = lines[crt_id + count].split(';')
            except:
                break
            
            if [new_line[1], new_line[2]] == [current_line[1], current_line[2]]: # The line has the same name and type
                # We append it because is a Option of the same Decision
                related_lines.append(new_line)
                count += 1
            else:
                break
        crt_id += count - 1 # We jump the related Options

        options = []
        for row in related_lines:
            # For each related Option, we prepare it and add to the options variable
            id = row[0]
            dtype = row[1]
            decision_name = row[2]
            option = row[3]
            dependencies = row[4]

            if 'E-' == dtype[0:2]:
                # If the Option is a ending, we add it to the endings variable
                ending_name = dtype[2:]
                if ending_name not in endings: endings[ending_name] = 0

            points = [p.strip() for p in row[5:5 + len(novel_points)]]
            points, dependencies = GetPointsAndDependencies(points, dependencies)

            options.append(Option(
                id = id,
                name = option,
                points = points,
                dependencies = dependencies
            ))

        # Once we have all related Options, create the Decision
        decisions.append(Decision(
            id=id,
            type=dtype,
            name=decision_name,
            options=options,
        ))

        crt_id+=1

    return decisions, novel_points, endings


def ConditionIsRight(left, operator, right):
    ''' Return True if (left operator right), otherwise return False '''
    result = False
    if right is not None:
        if operator == '<':
            if left < right:
                result = True
        if operator == '<=':
            if left <= right:
                result = True
        if operator == '>':
            if left > right:
                result = True
        if operator == '>=':
            if left >= right:
                result = True
        if operator == '=':
            if left == right:
                result = True

    return result

def GetSortedActionChain(ways:list[ActionChain]):
    '''
    Gets a list of ActionChain and returns it sorted by IdList
    This function actually has no use because the recursive decison
    tree explorer automatically orders the list of ActionChain
    '''
    decisions = dict()
    sortedDecisions = []
    for way in ways:
        # We create a dict with ['idList'] = ActionChain
        decisions[way.get_decision_sequency()] = way
    keys = list(decisions.keys())
    keys.sort() # Sorting the keys
    for d in keys:
        sortedDecisions.append(decisions[d])

    return sortedDecisions


def GetEndingStatistics(endings:dict, roads:int):
    ''' Return a dict with statistics of endings given '''
    result = {}
    for key, value in endings.items():
        percent = round((value * 100) / roads, 2)
        to_add = {
            'count': value,
            'percent': f'{percent:.2f}',
            'index': f'{(percent / 100):.2f}'
        }
        result[key] = to_add

    return result